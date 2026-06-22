"""Generate the Clinical Data Explorer risk assessment workbook.

Mirrors the structure/styling of docs_assets/Risk Assessment Template.xlsx
(Risk Management + Risk Matrix sheets) and populates it with the components
and failure scenarios of this project.
"""

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT = "docs_assets/Clinical Data Explorer - Risk Assessment.xlsx"

# ---- Risk matrix lookup (from the template's Risk Matrix sheet) ----------
# matrix[probability][severity] -> numeric risk index (1-4)
MATRIX = {
    "Critical": {"Low": 2, "Medium": 3, "High": 4, "Critical": 4},
    "High":     {"Low": 1, "Medium": 2, "High": 3, "Critical": 4},
    "Medium":   {"Low": 1, "Medium": 2, "High": 3, "Critical": 3},
    "Low":      {"Low": 1, "Medium": 1, "High": 2, "Critical": 3},
}


def ri(severity, probability):
    return MATRIX[probability][severity]


# ---- Styling -------------------------------------------------------------
ORANGE = "FFFF6543"          # template header accent
GROUP_FILL = "FFEFEFEF"      # template light grey (used on GxP col)
GROUP_BG = "FFF3F1FF"        # subtle purple group banding
WHITE = "FFFFFFFF"

INDEX_FILL = {
    1: "FFC6EFCE",  # green  - Low
    2: "FFFFEB9C",  # yellow - Medium
    3: "FFFFD199",  # orange - High
    4: "FFFFC7CE",  # red    - Critical
}

header_font = Font(name="Calibri", bold=True, color=WHITE, size=11)
group_font = Font(name="Calibri", bold=True, size=11)
base_font = Font(name="Calibri", size=10)

header_fill = PatternFill("solid", fgColor=ORANGE)
group_fill = PatternFill("solid", fgColor=GROUP_BG)

center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="top", wrap_text=True)
center_top = Alignment(horizontal="center", vertical="top", wrap_text=True)

thin = Side(style="thin", color="FFBFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ---- Column layout (matches the template's Example sheet) ----------------
HEADERS = [
    "PR #",                              # A
    "ID# : Title/Description",           # B
    "GxP Assessment - 21CFR11 Impact",   # C
    "Risk/Failure Scenarios",            # D
    "Severity",                          # E
    "Probability",                       # F
    "Risk Index",                        # G
    "GAMP5",                             # H
    "Risk Mitigation Steps",             # I
    "Verification of Mitigation Steps",  # J
    "S2",                                # K
    "P2",                                # L
    "RI2",                               # M
]
WIDTHS = [12.9, 33.8, 13.9, 40, 9.4, 11, 9.4, 7.5, 45.1, 40, 9.4, 9.4, 8]

GAMP5 = 5  # whole app is custom software

# Each group: (title, [scenario rows])
# scenario row: (component, scenario, sev, prob, mitigation, verification, s2, p2)
GROUPS = [
    ("Dataset Discovery & Selection", [
        ("Dataset Listing",
         "Dataset/snapshot list fails to load or omits available files, so users cannot find or select the correct clinical dataset.",
         "High", "Low",
         "Surface backend errors to the UI with retry; log Domino API failures to stdout; cover dataset listing in the e2e smoke test.",
         "Verify dataset dropdown populates against Domino datasets, snapshots, and NetApp volumes in e2e/integration tests.",
         "High", "Low"),
        ("Dataset Listing",
         "Wrong snapshot/revision is presented or selected, leading to analysis on an outdated or incorrect data version.",
         "High", "Medium",
         "Preserve and display snapshot identity (id + revision) through load; show selected version in the header.",
         "Test that snapshot id/revision shown matches the loaded file; confirm reload returns the same revision.",
         "High", "Low"),
        ("Source Resolution",
         "File source (Domino dataset, snapshot deeplink, or NetApp volume path) resolves to the wrong file, so users analyze unintended data.",
         "High", "Low",
         "Validate resolved file path/identity before handing to the MCP server; reject ambiguous resolutions.",
         "Add tests covering each source type; assert resolved path matches the requested source.",
         "High", "Low"),
    ]),
    ("Dataset Loading", [
        ("File Download",
         "File download fails or is truncated, producing an incomplete DataFrame and inaccurate downstream results.",
         "High", "Medium",
         "Verify downloaded byte size against source; fail the load on mismatch; enforce DATA_FILE_SIZE_LIMIT_B.",
         "Integrity check on download size; e2e load of a known file and row-count assertion.",
         "High", "Low"),
        ("File Parsing / DataFrame Conversion",
         "Parsing of SAS/XPT/CDISC Dataset-JSON misinterprets types, encodings, or special/missing-value codes, corrupting values.",
         "High", "Medium",
         "Use format-specific readers; preserve declared dtypes and missing-value semantics; validate against source metadata.",
         "Round-trip parsing tests per format with known fixtures comparing values, dtypes, and null counts.",
         "High", "Low"),
        ("File Parsing / DataFrame Conversion",
         "Numeric precision or date/datetime conversion errors during load alter clinical values.",
         "High", "Low",
         "Preserve numeric precision and timezone/date handling; avoid lossy float coercion of identifiers.",
         "Fixture tests asserting numeric precision and date values match source for representative columns.",
         "High", "Low"),
    ]),
    ("Dataset Load Queue & Memory Management", [
        ("Admission Queue",
         "Memory projection underestimates RAM use, causing pod OOM and loss of all in-flight sessions and loaded data.",
         "High", "Low",
         "Conservative DataFrame size estimate (~5x source); per-pod queue baseline + projection; size pod memory per tier.",
         "Load-test concurrent loads near the limit; confirm rejection rather than OOM.",
         "High", "Low"),
        ("Admission Queue",
         "Queue rejects valid load requests under load, making the app unavailable for legitimate analysis.",
         "Medium", "Medium",
         "Return a clear capacity error to the user; tune DATASET_LOAD_REQUEST_QUEUE_MAX_LENGTH; enable autoscaling with sticky sessions.",
         "Verify a clear capacity error is shown and a retry succeeds once capacity frees up.",
         "Medium", "Low"),
    ]),
    ("Table View & Column Labels", [
        ("Pagination & Sorting",
         "Pagination or sorting returns rows from the wrong offset or drops rows, misrepresenting the dataset contents.",
         "High", "Low",
         "Deterministic, stable sort/paging on the server; total-row count returned with each page.",
         "Contract tests asserting page boundaries, ordering, and totals against a fixture dataset.",
         "High", "Low"),
        ("Friendly / Column Labels",
         "Column label mapping maps a label to the wrong underlying variable, causing misinterpretation of clinical fields.",
         "High", "Low",
         "Map labels strictly by exact column_name; never reorder columns; show raw name on hover/toggle.",
         "Test that label toggle preserves column-to-variable mapping; spot-check CDISC ADaM labels.",
         "High", "Low"),
    ]),
    ("Filtering", [
        ("Standard Filters",
         "Filter operator logic (between, is/is not, contains, greater than) returns an incorrect subset, leading to wrong conclusions.",
         "High", "Medium",
         "Centralize and unit-test operator semantics; show active filters as badges; reflect filtered row count.",
         "Per-operator tests on fixtures comparing filtered counts and boundary values.",
         "High", "Low"),
        ("Missing-Value Filter",
         "Missing-value filter misclassifies nulls vs blanks vs special/sentinel codes, including or excluding wrong rows.",
         "Medium", "Medium",
         "Define a single canonical 'missing' definition shared by filters and missing-value analysis.",
         "Tests asserting consistent missing semantics across filter and analytics paths.",
         "Medium", "Low"),
    ]),
    ("Expression Filters (SAS WHERE / R dplyr / pandas)", [
        ("Expression Parser/Evaluator",
         "An expression is misparsed across dialects and silently returns the wrong rows instead of raising an error.",
         "High", "Medium",
         "Strict parsing with clear errors on unsupported syntax; never fall back to an empty/loose match silently.",
         "Sample-expression tests per dialect (see /table/expression_samples) comparing results to expected rows.",
         "High", "Low"),
        ("Expression Parser/Evaluator",
         "Arbitrary expression evaluation enables unsafe server-side code execution (injection) on the data pod.",
         "Critical", "Low",
         "Evaluate via a sandboxed/whitelisted expression engine; never eval untrusted strings; restrict to column/operator grammar.",
         "Security review of the evaluator; tests that injection payloads are rejected, not executed.",
         "Critical", "Low"),
    ]),
    ("Summary Statistics & Missing-Value Analysis", [
        ("Summary / Statistic Computation",
         "Aggregate statistics (mean, median, std, counts) are computed incorrectly or handle missing values inconsistently.",
         "High", "Medium",
         "Use vetted pandas aggregations; document missing-value handling; compute on the active filtered view.",
         "Compare computed stats to independently calculated reference values on fixtures.",
         "High", "Low"),
        ("Missing-Value Analysis",
         "Missing-value counts/badges undercount or overcount nulls, masking real data-quality issues.",
         "Medium", "Medium",
         "Reuse the canonical missing definition; show per-column and total counts; allow click-to-filter validation.",
         "Tests asserting badge counts equal actual null counts per column on fixtures.",
         "Medium", "Low"),
    ]),
    ("Charts & Visualization", [
        ("Histogram / X-Y Aggregation",
         "Binning or aggregation (avg/sum/count/min/max) is miscomputed, producing a misleading chart.",
         "Medium", "Medium",
         "Server-side aggregation reused from the stats layer; show n/mean/median/std with the chart.",
         "Tests comparing chart aggregates to table summary values for the same column/filter.",
         "Medium", "Low"),
        ("Chart Axis / Group Mapping",
         "Axis or group-by mapping is mismatched, so the chart misrepresents the relationship between variables.",
         "Medium", "Low",
         "Bind axis selections explicitly to columns; label axes with the selected variable.",
         "e2e checks that selected X/Y columns match rendered axis labels and data.",
         "Medium", "Low"),
    ]),
    ("AI Chat Agent (LLM)", [
        ("Natural-Language Query / Tool Calling",
         "The LLM hallucinates statistics or misinterprets the question, presenting fabricated results as factual.",
         "High", "High",
         "Ground answers in MCP tool results rather than free generation; show the underlying numbers/tool output; label AI responses.",
         "External (@external) chat tests asserting answers are backed by tool calls; spot-check for fabricated values.",
         "High", "Medium"),
        ("Data Exposure to LLM Provider",
         "Sensitive clinical data is sent to an external LLM provider, creating a data-exposure and compliance risk.",
         "High", "Medium",
         "Document data flow; support local/in-tenant models (Ollama/Azure); allow disabling chat; avoid sending raw PHI where possible.",
         "Confirm chat is opt-in via config; review what payloads are sent to the provider.",
         "High", "Low"),
        ("Chat Context / History Cache",
         "Chat history from a prior dataset or another session bleeds into the current answer.",
         "High", "Low",
         "Clear chat history on successful dataset switch; key history by session ID; enforce cache caps.",
         "Tests that a dataset switch clears history and that sessions cannot read each other's history.",
         "High", "Low"),
    ]),
    ("Domino Governance Integration", [
        ("Governance Bundle Detection",
         "A governed file is not detected, so reviewers believe data is ungoverned and the Create Finding action is hidden.",
         "High", "Low",
         "Match attachment overviews to the exact loaded file/revision; show explicit Governed/Not Governed badge.",
         "External tests that a known governed file shows the Governed badge and Create Finding.",
         "High", "Low"),
        ("Governance Bundle Detection",
         "A file is incorrectly shown as governed, giving false assurance of compliance tracking.",
         "Medium", "Low",
         "Require a positive bundle match before showing Governed; default to Not Governed on ambiguity.",
         "Tests asserting ungoverned files render the Not Governed badge.",
         "Medium", "Low"),
        ("Finding Submission",
         "Finding submission fails silently or is sent to the wrong bundle, losing the audit/compliance record.",
         "High", "Low",
         "Confirm submission success in the UI; bind finding to the resolved bundle id; surface API errors.",
         "External test creating a finding and verifying it appears on the correct bundle.",
         "High", "Low"),
        ("Finding Permalink Context",
         "The data-view permalink (active filters) is omitted or incorrect in the finding, so reviewers cannot reproduce the view.",
         "Medium", "Medium",
         "Append the current filtered-view URL to the finding description automatically; require deep-linking enabled.",
         "Test that a created finding contains a permalink that reproduces the same filtered view.",
         "Medium", "Low"),
    ]),
    ("Session & Identity", [
        ("Session Management",
         "Session ID collision or multi-tab use causes one browser session to see another session's dataset or results.",
         "High", "Low",
         "Use unguessable session IDs in a signed cookie; document the single-active-dataset-per-session limitation.",
         "Tests that distinct sessions cannot access each other's loaded dataset or chat.",
         "High", "Low"),
        ("Session Cookie Security",
         "An unsigned or insecure session cookie allows session tampering or fixation.",
         "High", "Low",
         "Sign cookies with a strong secret; set Secure/HttpOnly/SameSite; rotate the secret on deploy.",
         "Security review of cookie attributes; test tampered cookies are rejected.",
         "High", "Low"),
        ("Authorization / Identity Propagation",
         "App-owner permissions are used instead of the viewing user's, exposing datasets beyond the viewer's entitlement.",
         "High", "Medium",
         "Use viewer Bearer token via Domino Extension identity propagation; document the owner-permission limitation until enforced.",
         "Verify identity-propagated calls use the viewer token; review access scope per environment.",
         "High", "Low"),
    ]),
    ("Permalinks & Deep Links", [
        ("View State in URL",
         "A shared link reconstructs a different filter/view than intended, leading a reviewer to the wrong data.",
         "Medium", "Low",
         "Encode complete view state in the URL; validate on load; require 'Deep linking and query parameters' enabled.",
         "Tests that a copied link restores identical filters and resulting row set.",
         "Medium", "Low"),
    ]),
]


def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Risk Management"

    ncols = len(HEADERS)

    # Title band (row 1): Risk Assessment (A1:G1), Risk Control (H1:M1)
    ws.merge_cells("A1:G1")
    ws.merge_cells("H1:M1")
    a1 = ws["A1"]
    a1.value = "Risk Assessment"
    h1 = ws["H1"]
    h1.value = "Risk Control"
    for cell in (a1, h1):
        cell.fill = header_fill
        cell.font = Font(name="Calibri", bold=True, color=WHITE, size=12)
        cell.alignment = center
    ws.row_dimensions[1].height = 18

    # Header row (row 2)
    for i, head in enumerate(HEADERS, start=1):
        c = ws.cell(row=2, column=i, value=head)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
        c.border = border
    ws.row_dimensions[2].height = 42

    # Column widths
    for i, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A3"

    r = 3
    for title, scenarios in GROUPS:
        # group header row
        for col in range(1, ncols + 1):
            gc = ws.cell(row=r, column=col)
            gc.fill = group_fill
            gc.border = border
        ws.cell(row=r, column=2, value=title).font = group_font
        ws.cell(row=r, column=2).alignment = Alignment(
            horizontal="left", vertical="center", wrap_text=True)
        cgx = ws.cell(row=r, column=3, value="Yes")
        cgx.alignment = center
        cgx.font = group_font
        cg5 = ws.cell(row=r, column=8, value=GAMP5)
        cg5.alignment = center
        cg5.font = group_font
        ws.row_dimensions[r].height = 20
        r += 1

        for (comp, scenario, sev, prob, mit, ver, s2, p2) in scenarios:
            idx = ri(sev, prob)
            ridx2 = ri(s2, p2)
            row_vals = [
                "",            # A PR#
                comp,          # B
                "Yes",         # C
                scenario,      # D
                sev,           # E
                prob,          # F
                idx,           # G
                GAMP5,         # H
                mit,           # I
                ver,           # J
                s2,            # K
                p2,            # L
                ridx2,         # M
            ]
            for col, val in enumerate(row_vals, start=1):
                c = ws.cell(row=r, column=col, value=val)
                c.border = border
                c.font = base_font
                if col in (2, 4, 9, 10):           # text columns
                    c.alignment = left
                else:
                    c.alignment = center_top
            # color-code risk index cells
            ws.cell(row=r, column=7).fill = PatternFill("solid", fgColor=INDEX_FILL[idx])
            ws.cell(row=r, column=13).fill = PatternFill("solid", fgColor=INDEX_FILL[ridx2])
            ws.row_dimensions[r].height = 58
            r += 1

    last = r - 1
    # Data validations
    sev_dv = DataValidation(type="list", formula1='"Critical,High,Medium,Low"', allow_blank=True)
    idx_dv = DataValidation(type="list", formula1='"1,2,3,4"', allow_blank=True)
    ws.add_data_validation(sev_dv)
    ws.add_data_validation(idx_dv)
    sev_dv.add(f"E3:F{last}")
    sev_dv.add(f"K3:L{last}")
    idx_dv.add(f"G3:G{last}")
    idx_dv.add(f"M3:M{last}")

    # ---- Risk Matrix sheet (reference, copied from template) -------------
    build_matrix_sheet(wb)

    wb.save(OUT)
    print("Wrote", OUT, "with rows through", last)


def build_matrix_sheet(wb):
    ws = wb.create_sheet("Risk Matrix")
    title_fill = header_fill
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 70
    for col in "CDEF":
        ws.column_dimensions[col].width = 11
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 22

    def put(addr, val, bold=False, fill=None, align=center, color=None):
        c = ws[addr]
        c.value = val
        c.font = Font(name="Calibri", bold=bold, size=10, color=color)
        c.alignment = align
        if fill:
            c.fill = fill
        c.border = border

    # Matrix block
    put("A1", "Risk Index", bold=True, fill=title_fill, color=WHITE)
    ws.merge_cells("C1:F1")
    put("C1", "Severity", bold=True, fill=title_fill, color=WHITE)
    ws.merge_cells("H1:I1")
    put("H1", "GAMP5 Category", bold=True, fill=title_fill, color=WHITE)

    sev_order = ["Low", "Medium", "High", "Critical"]
    for j, s in enumerate(sev_order):
        put(f"{get_column_letter(3+j)}2", s, bold=True, fill=group_fill)
    put("H2", "Category", bold=True, fill=group_fill)
    put("I2", "Detail", bold=True, fill=group_fill)

    ws.merge_cells("A3:A6")
    put("A3", "Probability", bold=True, fill=group_fill)
    prob_order = ["Critical", "High", "Medium", "Low"]
    for i, p in enumerate(prob_order):
        rr = 3 + i
        put(f"B{rr}", p, bold=True, fill=group_fill)
        for j, s in enumerate(sev_order):
            val = MATRIX[p][s]
            put(f"{get_column_letter(3+j)}{rr}", val,
                fill=PatternFill("solid", fgColor=INDEX_FILL[val]))

    gamp = [(1, "Operating Systems"), (3, "Standard Software"),
            (4, "Configured Software"), (5, "Custom Software")]
    for i, (cat, det) in enumerate(gamp):
        rr = 3 + i
        put(f"H{rr}", cat)
        put(f"I{rr}", det, align=left)

    put("A8", "Classification:", bold=True)
    put("A10", "Risk Index", bold=True, fill=group_fill)
    put("B10", "Definition", bold=True, fill=group_fill)
    defs = [
        ("4 - Critical",
         "Severe impact on operations, safety, compliance, or legal standing; potentially catastrophic.\n\n"
         "Requires immediate action, comprehensive mitigation plans, and top management involvement to prevent or address the risk.\n\n"
         "Must be addressed before release."),
        ("3 - High",
         "Significant potential to disrupt operations, cause safety issues, or result in non-compliance.\n\n"
         "Demands prioritized attention, proactive mitigation strategies, and regular monitoring.\n\n"
         "Must be addressed before release."),
        ("2 - Medium",
         "Potential to disrupt operations or cause moderate compliance issues.\n\n"
         "Requires targeted monitoring and preemptive measures to reduce the likelihood or impact.\n\n"
         "Should be addressed before release, can be addressed after release if a justifiable workaround is available "
         "and if release delay would lead to significant other problems.\n\n"
         "Mitigation plans must be communicated to appropriate personnel (including users)."),
        ("1 - Low",
         "Minimal impact on operations, safety, or compliance; easily managed if it occurs.\n\n"
         "Routine monitoring and standard procedures are sufficient to mitigate this risk.\n\n"
         "Does not block release of software for its intended use.\n\n"
         "Can be addressed in a future release, justification must be provided."),
    ]
    for i, (label, text) in enumerate(defs):
        rr = 11 + i
        idx = int(label[0])
        put(f"A{rr}", label, bold=True, fill=PatternFill("solid", fgColor=INDEX_FILL[idx]))
        put(f"B{rr}", text, align=left)
        ws.row_dimensions[rr].height = 95


if __name__ == "__main__":
    main()
